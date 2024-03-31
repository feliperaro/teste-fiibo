"""
excel.py - A Python module for working with Excel files (.xlsx).

This module provides functions to write data to and read data from Excel files.
It uses the openpyxl library for handling Excel workbooks.
"""
import os
import openpyxl
import gvars


def read_xlsx(path: str) -> list[dict]:
    """
    Reads data from an XLSX file and outputs it as a dictionary.
    inputs:
        filename -> The path to the XLSX file.
    outputs: 
        data -> list of dictionaries
        dictionary containing the data from the XLSX file.
    """
    print("read_xlsx()")

    if not ".xlsx" in path:
        return -1

    if not os.path.exists(path):
        return -2

    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    data = []
    for index, row in enumerate(sheet.iter_rows()):
        if index == 0:
            continue

        data.append({
            "notas": row[0].value,
            "competencia": row[1].value,
            "status": row[2].value,
            "valor": row[3].value,
            "data_criacao": str(row[4].value),
        })

    return data


def test_read_xlsx():
    """
    Test read_xlsx() function
    """
    path = gvars.EXCEL_FILEPATH_TEST
    test_cases = [
        {
            "test_case_name": "Invalid file format",
            "path": "path/to/invalid.txt",
            "expected_result": -1,
        },
        {
            "test_case_name": "Non-existent path",
            "path": "path/non_existent.xlsx",
            "expected_result": -2,
        },
        {
            "test_case_name": "Valid XLSX file",
            "path": path,
            "expected_result": [
                {
                    "notas": "value1",
                    "competência": "value2", 
                    "status": "value3", 
                    "valor": 10.5, 
                    "data_criacao": "2023-01-01"
                },
            ],
        },
    ]

    for test_case in test_cases:
        test_case_name = test_case["test_case_name"]
        print("testing...", test_case_name)

        expected_result = test_case["expected_result"]
        print("expected_result", expected_result)

        result = read_xlsx(test_case["path"])
        print("result", result)

        assert result == expected_result
        print(f"test '{test_case_name}' success!")


def write_to_xlsx(data: list[dict], path: str):
    """
    Writes data to an XLSX file.
    inputs:
        data (list[dict]): A list of dictionaries containing the data to write.
        path (str): The path to the output XLSX file.
    """
    if not isinstance(data, list):
        return -1

    if not data or len(data) == 0:
        return 0

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    keys = data[0].keys()
    list_of_keys = list(keys)
    sheet.append(list_of_keys)
    for row in data:
        row_values = list(row.values())
        sheet.append(row_values)

    workbook.save(path)
    return 1


def test_write_to_xlsx():
    """
    Test write_to_xlsx() function
    """
    path = gvars.EXCEL_FILEPATH_TEST
    test_data = [
        {
            "notas": "value1",
            "competência": "value2", 
            "status": "value3", 
            "valor": 10.5, 
            "data_criacao": "2023-01-01"
        },
    ]

    test_cases = [
        {
            "test_case_name": "invalid_data",
            "data": "string is not valid data",
            "path": path,
            "expected_result": -1,
        },
        {
            "test_case_name": "empty_data",
            "data": [],
            "path": path,
            "expected_result": 0,
        },
        {
            "test_case_name": "Valid data and path",
            "data": test_data,
            "path": path,
            "expected_result": 1,
        },
    ]

    for test_case in test_cases:
        test_case_name = test_case["test_case_name"]
        print("test_case_name", test_case_name)

        expected_result = test_case["expected_result"]
        print("expected_result", expected_result)

        result = write_to_xlsx(test_case["data"], test_case["path"])
        print("result:", result)

        assert result == expected_result
        print(f"test '{test_case_name}' success!")


if __name__ == "__main__":
    test_write_to_xlsx()
    test_read_xlsx()
